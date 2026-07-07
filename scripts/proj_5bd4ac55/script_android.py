# coding:utf-8
import uiautomator2 as u2
import time
import datetime
import openpyxl
import os
from PIL import Image
import io


class Excle_wr:
    def __init__(self, excle_path):
        self.excle_path = excle_path
        # 判断文件是否存在，不存在则创建
        if not os.path.exists(excle_path):
            wb = openpyxl.Workbook()
            wb.save(excle_path)
        self.excle = openpyxl.load_workbook(excle_path)
        self.sheet = self.excle.active  # 实例化表

    def read_excle(self):
        maxrows = self.sheet.max_row  # 行
        maxcols = self.sheet.max_column  # 列
        data = {}
        sn_data = []
        # 取出表中的第一行作为字典的key
        for i in range(1, maxcols + 1):
            data.setdefault(self.sheet.cell(1, i).value)
        for i in range(2, maxrows + 1):
            for j in range(1, maxcols + 1):
                data[self.sheet.cell(1, j).value] = self.sheet.cell(i, j).value
            sn_data.append(data.copy())
        return sn_data

    def write_excle(self, row, col, w_data):
        try:
            self.sheet.cell(row, col).value = w_data
            self.excle.save(self.excle_path)
            return True
        except Exception as e:
            print(f"Excel写入失败：{e}")
            return False

    def close(self):
        # 关闭Excel文件，释放资源
        self.excle.close()


class FileOperations:
    def __init__(self, filepath):
        self.filepath = filepath
        # 确保日志目录存在
        os.makedirs(filepath, exist_ok=True)

    # 清空文件
    def vEmptyFile(self, strFileName="log.txt"):
        try:
            file_path = os.path.join(self.filepath, strFileName)
            with open(file_path, 'w+', encoding='utf-8') as fl:
                fl.write('')
            return True
        except Exception as e:
            print(f"清空文件失败：{e}")
            return False

    # 写文件并选择方式和是否打印写入信息
    def write_log(self, log_info="", logname="log.txt", writeway='a+', time_mark=True):
        try:
            # 用os.path.join拼接路径，兼容跨平台
            file_path = os.path.join(self.filepath, logname)
            with open(file_path, writeway, encoding='utf-8') as fl:
                time_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if time_mark:
                    log_content = f"[{time_str}]{str(log_info)}\n".replace("\r", "")
                else:
                    log_content = f"[{time_str}]{str(log_info)}"
                fl.write(log_content)
                print(log_content.strip())
            return True
        except Exception as e:
            print(f"写日志失败：{e}")
            return False


class T2P:
    def __init__(self, filepath, excle_filepath, screenshot_path,logname):
        self.logname = logname
        self.write_log = FileOperations(filepath)
        self.filepath = excle_filepath
        self.screenshot_path = screenshot_path
        # 确保截图目录存在
        os.makedirs(screenshot_path, exist_ok=True)
        self.start_times = 1
        self.excle_wr = None  # 延迟初始化Excel

    # 连接手机，打开app
    def conn_phone(self, phone_name, appname):
        try:
            # USB连接
            self.driver = u2.connect_usb(phone_name)
            self.driver.app_stop(appname)
            self.write_log.write_log(f"成功连接设备：{phone_name}", self.logname)
        except Exception as e:
            self.write_log.write_log(f"连接设备失败：{e}", self.logname)
            raise  # 连接失败直接终止，避免后续无效执行

    # 截屏
    def screenshot(self, times=""):
        try:
            screenshot_name = os.path.join(self.screenshot_path,
                f"{times}--{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            )
            self.driver.screenshot(screenshot_name)
            self.write_log.write_log(f"截图保存：{screenshot_name}", self.logname)
            return screenshot_name
        except Exception as e:
            self.write_log.write_log(f"截图失败：{e}", self.logname)
            return ""

    def is_black_screen(self, img):
        # 图片亮度值检查
        try:
            grayscale_img = img.convert('L')
            pixel_count = grayscale_img.width * grayscale_img.height
            if pixel_count == 0:
                return False
            avg_brightness = sum(grayscale_img.getdata()) / pixel_count
            self.write_log.write_log(f"计算亮度值：{round(avg_brightness, 2)}", self.logname)
            return round(avg_brightness, 2)
        except Exception as e:
            self.write_log.write_log(f"黑屏检测失败：{e}", self.logname)
            return False

    # 小米11 杀后台（可选）
    def kill_app(self):
        try:
            self.driver(resourceId='com.android.systemui:id/center_group').click()
            time.sleep(0.5)
            self.driver(resourceId="com.android.systemui:id/recent_apps").click()
            time.sleep(0.5)
            self.driver(resourceId="com.miui.home:id/clearAnimView").click()
            self.write_log.write_log("成功杀后台", self.logname)
        except Exception as e:
            self.write_log.write_log(f"杀后台失败：{e}", self.logname)

    def live_broadcast(self, display, appname, device_name, times):
        # 初始化Excel并写入表头
        self.excle_wr = Excle_wr(self.filepath)
        header = ["次数", "结果", "加载时间", "黑屏检查1", "黑屏检查2"]
        for col, h in enumerate(header, 1):
            self.excle_wr.write_excle(1, col, h)

        # 直播循环
        for i in range(self.start_times, times + 1):
            self.write_log.write_log(f"直播压测-------------------------------------第{i}次", self.logname)
            self.start_times = i
            try:
                # 打开APP
                self.write_log.write_log("打开APP", self.logname)
                self.driver.app_stop(appname)  # 先停止，避免多实例
                time.sleep(2)
                self.driver.app_start(appname)
                s_flag = time.time()

                # 等待设备页面加载
                if self.driver(text=device_name).exists(timeout=20):
                    e_flag = time.time()
                    load_time = round(e_flag - s_flag, 2)
                    self.write_log.write_log(f"打开APP到刷出页面时间:{load_time}S", self.logname)
                    time.sleep(5)

                    # 进入直播
                    self.write_log.write_log("进入直播", self.logname)

                    # self.write_log.write_log(f"点击{display}", self.logname)
                    self.driver(resourceId="com.lockin.loock:id/ivThumb").click()

                    # self.driver(text=display).click()
                    start_time = time.time()

                    # 等待直播加载（KB/s标识）
                    if self.driver(textContains="KB/s").exists(60):
                        conn_time = round(time.time() - start_time, 2)
                        self.write_log.write_log(f"进入直播成功,用时{conn_time}s", self.logname)
                        # 写入基础信息
                        self.excle_wr.write_excle(1 + i, 1, i)
                        self.excle_wr.write_excle(1 + i, 2, "直播成功")
                        self.excle_wr.write_excle(1 + i, 3, conn_time)

                        # 黑屏检测
                        time.sleep(5)
                        self.write_log.write_log("等待5s，开始截图检查黑屏", self.logname)
                        black_screen_results = []  # 存储检测结果
                        for x in range(2):
                            try:
                                # 定位直播画面元素
                                s = self.driver.xpath('//android.widget.ScrollView/android.view.ViewGroup[1]/android.view.ViewGroup[1]/android.view.ViewGroup[2]')
                                if not s.exists:
                                    self.write_log.write_log("未找到直播画面元素", self.logname)
                                    black_screen_results.append("元素未找到")
                                    continue
                                # 截图并检测
                                sreen = s.screenshot()
                                result = self.is_black_screen(sreen)
                                # 处理检测结果
                                if result is False:
                                    black_screen_results.append("检测失败")
                                elif result < 10:
                                    black_screen_results.append("黑屏")
                                    self.write_log.write_log(f"直播画面亮度值{result}---黑屏", self.logname)
                                    self.screenshot(f"第{i}次-黑屏{x + 1}")
                                    if x == 0:
                                        self.write_log.write_log("等待10s，进行第二次检查", self.logname)
                                        time.sleep(10)
                                else:
                                    black_screen_results.append(f"正常({result})")
                                    self.write_log.write_log(f"直播画面亮度值{result}---正常", self.logname)
                                    break
                            except Exception as e:
                                self.write_log.write_log(f"第{x + 1}次黑屏检测异常：{e}", self.logname)
                                black_screen_results.append("异常")

                        # 写入黑屏检测结果
                        self.excle_wr.write_excle(1 + i, 4,
                                                  black_screen_results[0] if len(black_screen_results) > 0 else "无")
                        self.excle_wr.write_excle(1 + i, 5,
                                                  black_screen_results[1] if len(black_screen_results) > 1 else "无")

                    else:
                        # 直播失败
                        self.write_log.write_log("直播失败（未找到KB/s标识）", self.logname)
                        self.excle_wr.write_excle(1 + i, 1, i)
                        self.excle_wr.write_excle(1 + i, 2, "直播失败")
                        self.excle_wr.write_excle(1 + i, 3, datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
                        self.screenshot(f"第{i}次-直播失败")

                    # 清理环境
                    self.write_log.write_log("杀掉app，等待40s", self.logname)
                    self.driver.app_stop(appname)
                    time.sleep(40)

                else:
                    # 未进入设备页面
                    self.screenshot(f"第{i}次-未进入设备页面")
                    # self.kill_app()
                    self.driver.app_stop(appname)
                    self.write_log.write_log("未进入设备页面，杀掉app", self.logname)
                    time.sleep(5)

            except Exception as E:
                # 记录异常后继续下一次
                self.write_log.write_log(f"第{i}次压测出现异常：{E}", self.logname)
                self.screenshot(f"第{i}次-异常")
                self.driver.app_stop(appname)
                # 异常时写入Excel
                self.excle_wr.write_excle(1 + i, 1, i)
                self.excle_wr.write_excle(1 + i, 2, "异常")
                self.excle_wr.write_excle(1 + i, 3, str(E)[:50])  # 截断过长异常信息
                time.sleep(10)
                continue  # 继续下一次

        # 所有压测完成后关闭Excel
        if self.excle_wr:
            self.excle_wr.close()
        self.write_log.write_log(f"压测完成！共执行{times}次", self.logname)


if __name__ == "__main__":
    # 测试前配置
    now_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    appname = "com.lockin.loock"
    phone_name = ""  # USB设备名/WiFi地址
    device_name = "哈哈哈"
    display = "外屏"  # 601内屏/外屏，其他项目不涉及不管

    # 路径配置（用os.path.join，兼容跨平台）
    current_path = os.getcwd()
    result_path = os.path.join(current_path, "log",f"{device_name}直播{display}压测_保持时长{now_time}")
    screenshot_path = os.path.join(result_path, "截图")
    logname = f"直播压测_保持时长{now_time}.txt"
    log_path = result_path
    excle_filepath = os.path.join(result_path, f"直播{display}压测_保持时长{now_time}.xlsx")

    # 创建目录（exist_ok=True避免重复创建报错）
    os.makedirs(result_path, exist_ok=True)
    os.makedirs(screenshot_path, exist_ok=True)

    # 测试执行
    test_times = 5000  # 压测次数
    try:
        t2p = T2P(log_path, excle_filepath, screenshot_path,logname)
        t2p.conn_phone(phone_name, appname)
        t2p.live_broadcast(display, appname, device_name, test_times)
    except Exception as e:
        print(f"程序入口异常：{e}")